 Ejercicios Modulo 5. Excel(Enunciado)

 EXCEL

1) Ejercicio de manejo de Excel en base al fichero 02_Excel_data.xlsx:
* Abre el fichero Excel
* Lista las hojas que hay
* Crea una nueva hoja que se llame olimpiadas entre ventas y otros
* Vuelve a listar las hojas que hay
* Guarda los cambios y comprueba en el excel si se han efectuado  

import openpyxl

# Ruta al archivo Excel
archivo_excel = "02_Excel_data.xlsx"

# Abre el archivo Excel
workbook = openpyxl.load_workbook(archivo_excel)

# Lista las hojas que hay
hojas_existentes = workbook.sheetnames
print("Hojas existentes:", hojas_existentes)

# Crea una nueva hoja llamada "olimpiadas" entre "ventas" y "otros"
nueva_hoja = workbook.create_sheet("olimpiadas", index=hojas_existentes.index("otros"))

# Vuelve a listar las hojas que hay
hojas_actualizadas = workbook.sheetnames
print("Hojas después de añadir 'olimpiadas':", hojas_actualizadas)

# Guarda los cambios
workbook.save(archivo_excel)

# Cierra el archivo
workbook.close()

print("Cambios guardados. Comprueba el archivo Excel para verificar.")


2) Ejercicio:

import openpyxl

# Ruta al archivo Excel
archivo_excel = "02_Excel_data.xlsx"

# Abre el archivo Excel
workbook = openpyxl.load_workbook(archivo_excel)

# Selecciona la hoja "olimpiadas"
hoja_olimpiadas = workbook["olimpiadas"]

# Crea una lista con los datos olímpicos
datos_olimpicos = [
    ["USA", 46, 12, 5],
    ["China", 38, 20, 7],
    ["UK", 29, 7, 7],
    ["Russia", 22, 10, 9],
    ["South Korea", 13, 3, 2],
    ["Germany", 11, 7, 4]
]

# Añade la lista a la hoja "olimpiadas"
for fila in datos_olimpicos:
    hoja_olimpiadas.append(fila)

# Mover el rango de datos una fila hacia abajo
hoja_olimpiadas.move_range("A2:D7", rows=1, translate=True)

# Añadir cabeceras en la primera línea
cabeceras = ["Pais", "Oros", "Platas", "Bronces"]
hoja_olimpiadas.insert_rows(1)
for col, header in enumerate(cabeceras, start=1):
    hoja_olimpiadas.cell(row=1, column=col, value=header)

# Guardar los cambios
workbook.save(archivo_excel)

# Cerrar el archivo
workbook.close()

print("Cambios realizados y guardados. Comprueba el archivo Excel para verificar.")

3) Ejercicio Doc:( https://openpyxl.readthedocs.io/en/stable/styles.html):

import openpyxl
from openpyxl.styles import Font

# Ruta al archivo Excel
archivo_excel = "02_Excel_data.xlsx"

# Abre el archivo Excel
workbook = openpyxl.load_workbook(archivo_excel)

# Selecciona la hoja "olimpiadas"
hoja_olimpiadas = workbook["olimpiadas"]

# Aplicar estilo negrita a las cabeceras
cabeceras = ["Pais", "Oros", "Platas", "Bronces"]
bold_font = Font(bold=True)

for col, header in enumerate(cabeceras, start=1):
    hoja_olimpiadas.cell(row=1, column=col, value=header).font = bold_font

# Guardar los cambios
workbook.save(archivo_excel)

# Cerrar el archivo
workbook.close()

print("Cambios realizados y guardados. Comprueba el archivo Excel para verificar.")

4) Ejercicio:

import openpyxl
from openpyxl.styles import Font

# Ruta al archivo Excel
archivo_excel = "02_Excel_data.xlsx"

# Abre el archivo Excel
workbook = openpyxl.load_workbook(archivo_excel)

# Selecciona la hoja "olimpiadas"
hoja_olimpiadas = workbook["olimpiadas"]

# Añadir columna de sumatorio de medallas
for fila in hoja_olimpiadas.iter_rows(min_row=2, max_row=hoja_olimpiadas.max_row, min_col=2, max_col=4):
    suma_medallas = sum(cell.value if cell.value is not None else 0 for cell in fila)
    hoja_olimpiadas.cell(row=fila[0].row, column=5, value=suma_medallas)

# Guardar los cambios
workbook.save(archivo_excel)

# Cierra el archivo
workbook.close()

print("Cambios realizados y guardados. Comprueba el archivo Excel para verificar.")

5) Ejercicio:

import openpyxl
from openpyxl.chart import BarChart, Reference

# Ruta al archivo Excel
archivo_excel = "02_Excel_data.xlsx"

# Abre el archivo Excel
workbook = openpyxl.load_workbook(archivo_excel)

# Selecciona la hoja "olimpiadas"
hoja_olimpiadas = workbook["olimpiadas"]

# Crear un objeto BarChart
grafico_barras = BarChart()

# Configurar el título del gráfico
grafico_barras.title = "Medallas Olímpicas"

# Configurar categorías y datos
categorias = Reference(hoja_olimpiadas, min_col=1, min_row=2, max_row=hoja_olimpiadas.max_row)
datos = Reference(hoja_olimpiadas, min_col=2, min_row=1, max_col=5, max_row=hoja_olimpiadas.max_row)
grafico_barras.set_categories(categorias)
grafico_barras.add_data(datos, titles_from_data=True)

# Agregar el gráfico a la hoja
hoja_olimpiadas.add_chart(grafico_barras, "F2")

# Guardar los cambios
workbook.save(archivo_excel)



